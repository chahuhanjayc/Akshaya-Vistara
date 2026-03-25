"""
reports/views.py
"""

import io
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

from . import utils


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_dates(request):
    """Extract start_date / end_date from GET params with FY-start defaults."""
    today = date.today()
    default_start = today.replace(month=4, day=1) if today.month >= 4 \
                    else today.replace(year=today.year - 1, month=4, day=1)
    default_end = today

    try:
        start_date = date.fromisoformat(request.GET.get("start_date", ""))
    except ValueError:
        start_date = default_start

    try:
        end_date = date.fromisoformat(request.GET.get("end_date", ""))
    except ValueError:
        end_date = default_end

    return start_date, end_date


def _parse_as_of(request):
    """Extract as_of_date from GET params, defaulting to today."""
    try:
        return date.fromisoformat(request.GET.get("as_of_date", ""))
    except ValueError:
        return date.today()


def _xl_header(ws, company_name, title, subtitle, styles):
    """Write the standard 3-row company header into a worksheet, returns next row."""
    from openpyxl.styles import Alignment

    hdr_font   = styles["hdr_font"]
    title_font = styles["title_font"]
    sub_font   = styles["sub_font"]
    hdr_fill   = styles["hdr_fill"]

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value, c.font, c.fill = company_name, hdr_font, hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value, c.font = title, title_font
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value, c.font = subtitle, sub_font
    c.alignment = Alignment(horizontal="center")

    return 5  # data starts at row 5


# ─────────────────────────────────────────────────────────────────────────────
# REPORT VIEWS
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def reports_home(request):
    return render(request, "reports/reports_home.html")


@login_required
def profit_loss(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_profit_loss(company, start_date, end_date)
    return render(request, "reports/profit_loss.html", {
        "start_date": start_date,
        "end_date":   end_date,
        **data,
    })


@login_required
def balance_sheet(request):
    company = request.current_company
    as_of_date = _parse_as_of(request)
    data = utils.get_balance_sheet(company, as_of_date)
    return render(request, "reports/balance_sheet.html", {
        "as_of_date": as_of_date,
        **data,
    })


@login_required
def receivables_aging(request):
    company = request.current_company
    as_of_date = _parse_as_of(request)
    data = utils.get_receivables_aging(company, as_of_date)
    return render(request, "reports/receivables_aging.html", {
        "as_of_date": as_of_date,
        **data,
    })


@login_required
def trial_balance(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_trial_balance(company, start_date, end_date)
    return render(request, "reports/trial_balance.html", {
        "start_date": start_date,
        "end_date":   end_date,
        **data,
    })


@login_required
def gst_report(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_gst_report(company, start_date, end_date)
    return render(request, "reports/gst_report.html", {
        "start_date": start_date,
        "end_date":   end_date,
        **data,
    })


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORTS  (WeasyPrint — server-side, downloadable)
# ─────────────────────────────────────────────────────────────────────────────

def _render_report_pdf(request, template_name: str, context: dict, filename: str) -> HttpResponse:
    """
    Render a Django template to a WeasyPrint PDF and return as HttpResponse.
    context must already contain all template variables including current_company.
    """
    import weasyprint
    from django.template.loader import render_to_string

    html_str = render_to_string(template_name, context, request=request)
    pdf_bytes = weasyprint.HTML(string=html_str, base_url=request.build_absolute_uri("/")).write_pdf()
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def profit_loss_pdf(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_profit_loss(company, start_date, end_date)
    filename = (
        f"ProfitLoss_{company.name}_{start_date:%Y%m%d}_{end_date:%Y%m%d}.pdf"
        .replace(" ", "_")
    )
    return _render_report_pdf(
        request,
        "reports/profit_loss_pdf.html",
        {"current_company": company, "start_date": start_date, "end_date": end_date, **data},
        filename,
    )


@login_required
def balance_sheet_pdf(request):
    company = request.current_company
    as_of_date = _parse_as_of(request)
    data = utils.get_balance_sheet(company, as_of_date)
    filename = f"BalanceSheet_{company.name}_{as_of_date:%Y%m%d}.pdf".replace(" ", "_")
    return _render_report_pdf(
        request,
        "reports/balance_sheet_pdf.html",
        {"current_company": company, "as_of_date": as_of_date, **data},
        filename,
    )


@login_required
def trial_balance_pdf(request):
    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_trial_balance(company, start_date, end_date)
    filename = (
        f"TrialBalance_{company.name}_{start_date:%Y%m%d}_{end_date:%Y%m%d}.pdf"
        .replace(" ", "_")
    )
    return _render_report_pdf(
        request,
        "reports/trial_balance_pdf.html",
        {"current_company": company, "start_date": start_date, "end_date": end_date, **data},
        filename,
    )


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORTS
# ─────────────────────────────────────────────────────────────────────────────

def _xl_styles():
    """Return a dict of commonly used openpyxl styles."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        "hdr_font":   Font(name="Calibri", bold=True, size=14, color="FFFFFF"),
        "title_font": Font(name="Calibri", bold=True, size=12),
        "sub_font":   Font(name="Calibri", italic=True, size=10, color="555555"),
        "col_font":   Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "data_font":  Font(name="Calibri", size=10),
        "total_font": Font(name="Calibri", bold=True, size=10),
        "hdr_fill":   PatternFill("solid", fgColor="1F3864"),
        "col_fill":   PatternFill("solid", fgColor="2F5496"),
        "tot_fill":   PatternFill("solid", fgColor="E2EFDA"),
        "alt_fill":   PatternFill("solid", fgColor="F5F5F5"),
        "border":     border,
        "center":     Alignment(horizontal="center", vertical="center"),
        "right":      Alignment(horizontal="right"),
        "left":       Alignment(horizontal="left"),
        "num_fmt":    '#,##0.00',
    }


@login_required
def export_pl_excel(request):
    """Download Profit & Loss as a formatted .xlsx file."""
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font

    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_profit_loss(company, start_date, end_date)
    st = _xl_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    ws.sheet_view.showGridLines = False

    subtitle = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    row = _xl_header(ws, str(company.name), "Profit & Loss Statement", subtitle, st)

    # ── Income section ───────────────────────────────────────────────────────
    ws.cell(row=row, column=1).value = "INCOME"
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color="1A7942")
    row += 1

    ws.cell(row=row, column=1).value = "Ledger Account"
    ws.cell(row=row, column=2).value = "Amount (₹)"
    for col in (1, 2):
        c = ws.cell(row=row, column=col)
        c.font = st["col_font"]
        c.fill = PatternFill("solid", fgColor="1A7942")
        c.alignment = st["center"]
        c.border = st["border"]
    row += 1

    for i, item in enumerate(data["income_items"]):
        fill = st["alt_fill"] if i % 2 == 1 else None
        c1 = ws.cell(row=row, column=1, value=item["name"])
        c2 = ws.cell(row=row, column=2, value=float(item["amount"]))
        c1.font = st["data_font"]
        c2.font = st["data_font"]
        c2.number_format = st["num_fmt"]
        c2.alignment = st["right"]
        c1.border = c2.border = st["border"]
        if fill:
            c1.fill = c2.fill = fill
        row += 1

    # Total income row
    c1 = ws.cell(row=row, column=1, value="Total Income")
    c2 = ws.cell(row=row, column=2, value=float(data["total_income"]))
    c1.font = c2.font = st["total_font"]
    c1.fill = c2.fill = PatternFill("solid", fgColor="D5F0DC")
    c2.number_format = st["num_fmt"]
    c2.alignment = st["right"]
    c1.border = c2.border = st["border"]
    row += 2  # blank line

    # ── Expense section ──────────────────────────────────────────────────────
    ws.cell(row=row, column=1).value = "EXPENSES"
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11, color="CC0000")
    row += 1

    ws.cell(row=row, column=1).value = "Ledger Account"
    ws.cell(row=row, column=2).value = "Amount (₹)"
    for col in (1, 2):
        c = ws.cell(row=row, column=col)
        c.font = st["col_font"]
        c.fill = PatternFill("solid", fgColor="CC0000")
        c.alignment = st["center"]
        c.border = st["border"]
    row += 1

    for i, item in enumerate(data["expense_items"]):
        fill = st["alt_fill"] if i % 2 == 1 else None
        c1 = ws.cell(row=row, column=1, value=item["name"])
        c2 = ws.cell(row=row, column=2, value=float(item["amount"]))
        c1.font = st["data_font"]
        c2.font = st["data_font"]
        c2.number_format = st["num_fmt"]
        c2.alignment = st["right"]
        c1.border = c2.border = st["border"]
        if fill:
            c1.fill = c2.fill = fill
        row += 1

    c1 = ws.cell(row=row, column=1, value="Total Expenses")
    c2 = ws.cell(row=row, column=2, value=float(data["total_expense"]))
    c1.font = c2.font = st["total_font"]
    c1.fill = c2.fill = PatternFill("solid", fgColor="FFE0E0")
    c2.number_format = st["num_fmt"]
    c2.alignment = st["right"]
    c1.border = c2.border = st["border"]
    row += 2

    # ── Net Profit ───────────────────────────────────────────────────────────
    net = data["net_profit"]
    c1 = ws.cell(row=row, column=1, value="NET PROFIT" if net >= 0 else "NET LOSS")
    c2 = ws.cell(row=row, column=2, value=float(net))
    profit_color = "1F6B37" if net >= 0 else "CC0000"
    c1.font = c2.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c1.fill = c2.fill = PatternFill("solid", fgColor=profit_color)
    c2.number_format = st["num_fmt"]
    c2.alignment = st["right"]
    c1.border = c2.border = st["border"]

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"PL_{start_date}_{end_date}.xlsx"
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@login_required
def export_bs_excel(request):
    """Download Balance Sheet as a formatted .xlsx file."""
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font

    company = request.current_company
    as_of_date = _parse_as_of(request)
    data = utils.get_balance_sheet(company, as_of_date)
    st = _xl_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.sheet_view.showGridLines = False

    subtitle = f"As of {as_of_date.strftime('%d %b %Y')}"
    row = _xl_header(ws, str(company.name), "Balance Sheet", subtitle, st)

    def _write_section(label, items, total_label, total_val, header_color, total_color):
        nonlocal row
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11,
                                               color=header_color)
        row += 1
        ws.cell(row=row, column=1).value = "Ledger Account"
        ws.cell(row=row, column=2).value = "Balance (₹)"
        for col in (1, 2):
            c = ws.cell(row=row, column=col)
            c.font = st["col_font"]
            c.fill = PatternFill("solid", fgColor=header_color)
            c.alignment = st["center"]
            c.border = st["border"]
        row += 1

        for i, item in enumerate(items):
            fill = st["alt_fill"] if i % 2 == 1 else None
            c1 = ws.cell(row=row, column=1, value=item["name"])
            c2 = ws.cell(row=row, column=2, value=float(item["balance"]))
            c1.font = c2.font = st["data_font"]
            c2.number_format = st["num_fmt"]
            c2.alignment = st["right"]
            c1.border = c2.border = st["border"]
            if fill:
                c1.fill = c2.fill = fill
            row += 1

        c1 = ws.cell(row=row, column=1, value=total_label)
        c2 = ws.cell(row=row, column=2, value=float(total_val))
        c1.font = c2.font = st["total_font"]
        c1.fill = c2.fill = PatternFill("solid", fgColor=total_color)
        c2.number_format = st["num_fmt"]
        c2.alignment = st["right"]
        c1.border = c2.border = st["border"]
        row += 2

    _write_section("ASSETS", data["asset_items"], "Total Assets",
                   data["total_assets"], "1F497D", "DAEEF3")
    _write_section("LIABILITIES & EQUITY", data["liability_items"],
                   "Total Liabilities & Equity", data["total_liabilities"],
                   "7030A0", "EAD1DC")

    # Balance check row
    balanced = data["difference"] == 0
    msg = "✔ Books are BALANCED" if balanced else f"⚠ Out of balance by ₹{data['difference']:.2f}"
    c = ws.cell(row=row, column=1, value=msg)
    c.font = Font(name="Calibri", bold=True, size=10,
                  color="1F6B37" if balanced else "CC0000")
    ws.merge_cells(f"A{row}:B{row}")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"BalanceSheet_{as_of_date}.xlsx"
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


@login_required
def export_tb_excel(request):
    """Download Trial Balance as a formatted .xlsx file."""
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font

    company = request.current_company
    start_date, end_date = _parse_dates(request)
    data = utils.get_trial_balance(company, start_date, end_date)
    st = _xl_styles()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.sheet_view.showGridLines = False

    subtitle = f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    row = _xl_header(ws, str(company.name), "Trial Balance", subtitle, st)

    # Column headers
    headers = ["Ledger Account", "Group",
               "Opening Dr (₹)", "Opening Cr (₹)",
               "Period Dr (₹)", "Period Cr (₹)",
               "Closing Dr (₹)", "Closing Cr (₹)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = st["col_font"]
        c.fill = st["col_fill"]
        c.alignment = st["center"]
        c.border = st["border"]
    row += 1

    for i, r in enumerate(data["rows"]):
        fill = st["alt_fill"] if i % 2 == 1 else None
        vals = [r["name"], r["group"],
                float(r["opening_dr"]), float(r["opening_cr"]),
                float(r["period_dr"]),  float(r["period_cr"]),
                float(r["closing_dr"]), float(r["closing_cr"])]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = st["data_font"]
            c.border = st["border"]
            if col > 2:
                c.number_format = st["num_fmt"]
                c.alignment = st["right"]
            if fill:
                c.fill = fill
        row += 1

    # Totals row
    totals = ["TOTALS", "",
              float(data["tot_open_dr"]), float(data["tot_open_cr"]),
              float(data["tot_per_dr"]),  float(data["tot_per_cr"]),
              float(data["tot_clos_dr"]), float(data["tot_clos_cr"])]
    for col, v in enumerate(totals, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = st["total_font"]
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.border = st["border"]
        if col > 2:
            c.number_format = st["num_fmt"]
            c.alignment = st["right"]

    col_widths = [42, 12, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"TrialBalance_{start_date}_{end_date}.xlsx"
    resp = HttpResponse(buf, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp
